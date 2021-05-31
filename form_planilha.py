from openpyxl import Workbook
from openpyxl import load_workbook
arquivo_excel = Workbook()
wb = load_workbook('X.xlsx') #arquivo .xlsx
ws = wb['Folha1'] #página
for line in ws:    
	line[1].value = (str (line[1].value)).replace("+55","").replace("-", "").replace("(","").replace(")","").replace(" ","").replace(".","") #linha e oq deseja remover 

	print (line[1].value)
wb.save(filename = 'Teste.xlsx')

